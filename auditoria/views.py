from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime, timedelta
from .models import AuditLog
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import logging

audit_logger = logging.getLogger('audit')


class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = AuditLog.objects.all()
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.perfil.tipo_usuario in ['bibliotecario', 'dba']:
            return AuditLog.objects.all()
        return AuditLog.objects.filter(user=user)

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        queryset = self.get_queryset()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Logs de Auditoría"

        headers = ['Fecha', 'Usuario', 'Acción', 'Tipo Objeto', 'ID Objeto', 'Objeto', 'Cambios', 'IP']
        ws.append(headers)

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for log in queryset:
            ws.append([
                log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                log.user.username if log.user else 'N/A',
                log.get_action_display(),
                log.object_type,
                log.object_id or '',
                log.object_repr,
                str(log.changes),
                log.ip_address or '',
            ])

        for column in ws.columns:
            max_length = max((len(str(cell.value or '')) for cell in column), default=0)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename=audit_logs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        wb.save(response)
        return response

    @action(detail=False, methods=['get'])
    def statistics(self, request):
        queryset = self.get_queryset()
        thirty_days_ago = timezone.now() - timedelta(days=30)
        recent_logs = queryset.filter(timestamp__gte=thirty_days_ago)

        stats = {
            'total_logs': queryset.count(),
            'recent_logs': recent_logs.count(),
            'actions_by_type': {},
            'objects_by_type': {},
            'top_users': {},
        }

        for log in recent_logs:
            action_label = log.get_action_display()
            stats['actions_by_type'][action_label] = stats['actions_by_type'].get(action_label, 0) + 1
            stats['objects_by_type'][log.object_type] = stats['objects_by_type'].get(log.object_type, 0) + 1
            if log.user:
                stats['top_users'][log.user.username] = stats['top_users'].get(log.user.username, 0) + 1

        return Response(stats)


class LoginLogoutAPIView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request, *args, **kwargs):
        action = request.data.get('action')
        user = request.user if request.user.is_authenticated else None

        if action in ('login', 'logout') and user:
            AuditLog.objects.create(
                user=user,
                action=action.upper(),
                object_type='User',
                object_id=user.pk,
                object_repr=str(user),
                ip_address=self._get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', ''),
                changes={f'{action}_success': True},
            )

        return Response({'status': 'success'})

    def _get_client_ip(self, request):
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            return x_forwarded_for.split(',')[0].strip()
        return request.META.get('REMOTE_ADDR')
